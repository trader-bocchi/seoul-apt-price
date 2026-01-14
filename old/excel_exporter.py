"""
엑셀 파일 저장 모듈
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List
from datetime import datetime
from data_collector import Property, Complex


class ExcelExporter:
    """엑셀 파일 내보내기"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
    
    def export(
        self,
        filename: str,
        properties: List[Property],
        complexes: List[Complex],
        region_name: str,
        include_properties: bool = True,
        include_complexes: bool = True
    ):
        """
        엑셀 파일로 저장
        
        Args:
            filename: 저장할 파일명
            properties: 매물 리스트
            complexes: 단지 리스트
            region_name: 지역명
            include_properties: 매물 정보 포함 여부
            include_complexes: 단지 정보 포함 여부
        """
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 매물 정보 시트
            if include_properties and properties:
                df_properties = self._create_properties_dataframe(properties)
                df_properties.to_excel(writer, sheet_name='매물 정보', index=False)
                self._format_sheet(writer.book['매물 정보'])
            
            # 단지 정보 시트
            if include_complexes and complexes:
                df_complexes = self._create_complexes_dataframe(complexes)
                df_complexes.to_excel(writer, sheet_name='단지 정보', index=False)
                self._format_sheet(writer.book['단지 정보'])
            
            # 통계 정보 시트
            df_stats = self._create_stats_dataframe(properties, complexes, region_name)
            df_stats.to_excel(writer, sheet_name='통계 정보', index=False)
            self._format_sheet(writer.book['통계 정보'])
    
    def _create_properties_dataframe(self, properties: List[Property]) -> pd.DataFrame:
        """매물 정보 DataFrame 생성"""
        data = {
            '매물 ID': [p.item_id for p in properties],
            '지역명': [p.region_name for p in properties],
            '단지명': [p.complex_name for p in properties],
            '부동산 유형': [p.property_type for p in properties],
            '거래 유형': [p.trade_type for p in properties],
            '가격(만원)': [p.price for p in properties],
            '가격 표시': [p.price_display for p in properties],
            '위도': [p.latitude for p in properties],
            '경도': [p.longitude for p in properties],
            '관리비 최소': [p.min_mvi_fee for p in properties],
            '관리비 최대': [p.max_mvi_fee for p in properties],
            'VR 투어': ['예' if p.tour_exist else '아니오' for p in properties],
            '수집 일시': [p.collected_at.strftime('%Y-%m-%d %H:%M:%S') for p in properties]
        }
        return pd.DataFrame(data)
    
    def _create_complexes_dataframe(self, complexes: List[Complex]) -> pd.DataFrame:
        """단지 정보 DataFrame 생성"""
        data = {
            '단지 ID': [c.item_id for c in complexes],
            '단지명': [c.complex_name for c in complexes],
            '위도': [c.latitude for c in complexes],
            '경도': [c.longitude for c in complexes],
            '매매 중위 평당가(만원)': [c.deal_median_unit_price for c in complexes],
            '전세 중위 보증금(만원)': [c.lease_median_rate for c in complexes],
            '건축년도': [c.build_year for c in complexes],
            'VR 투어': ['예' if c.tour_exist else '아니오' for c in complexes],
            '매물 개수': [c.article_count for c in complexes]
        }
        return pd.DataFrame(data)
    
    def _create_stats_dataframe(self, properties: List[Property], complexes: List[Complex], region_name: str) -> pd.DataFrame:
        """통계 정보 DataFrame 생성"""
        now = datetime.now()
        
        # 거래 유형별 개수
        trade_counts = {}
        for prop in properties:
            trade_type = prop.trade_type
            trade_counts[trade_type] = trade_counts.get(trade_type, 0) + 1
        
        # 부동산 유형별 개수
        property_counts = {}
        for prop in properties:
            prop_type = prop.property_type
            property_counts[prop_type] = property_counts.get(prop_type, 0) + 1
        
        # 가격 통계
        prices = [p.price for p in properties if p.price > 0]
        avg_price = sum(prices) / len(prices) if prices else 0
        max_price = max(prices) if prices else 0
        min_price = min(prices) if prices else 0
        
        data = {
            '항목': [
                '수집 일시',
                '지역명',
                '총 매물 개수',
                '총 단지 개수',
                '매매 개수',
                '전세 개수',
                '월세 개수',
                '평균 가격(만원)',
                '최고가(만원)',
                '최저가(만원)'
            ],
            '값': [
                now.strftime('%Y-%m-%d %H:%M:%S'),
                region_name,
                len(properties),
                len(complexes),
                trade_counts.get('매매', 0),
                trade_counts.get('전세', 0),
                trade_counts.get('월세', 0),
                f"{avg_price:,.0f}" if avg_price > 0 else "0",
                f"{max_price:,.0f}" if max_price > 0 else "0",
                f"{min_price:,.0f}" if min_price > 0 else "0"
            ]
        }
        return pd.DataFrame(data)
    
    def _format_sheet(self, sheet):
        """시트 포맷팅"""
        # 헤더 행 포맷팅
        for cell in sheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 컬럼 너비 자동 조정
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

